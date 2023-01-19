import xlsxwriter
import openpyxl

from data.field_data import *
import bot.parcing_links


def Make_Table():
    try:
        f = open('config/data.xlsx')
        print('working on an existing table')
        f.close()
    except IOError:
        print('create a table')
        # Making xlsx file
        workbook = xlsxwriter.Workbook('config/data.xlsx')

        # offers
        offers = workbook.add_worksheet('offers')
        row = 0
        col = 0
        for field in main_fields:
            offers.write(row, col, field)
            col += 1

        # child offers
        child_offers = workbook.add_worksheet('child_offers')
        row = 0
        col = 0
        for field in child_fields:
            child_offers.write(row, col, field)
            col += 1

        # images
        images = workbook.add_worksheet('images')
        row = 0
        col = 0
        for field in image:
            images.write(row, col, field)
            col += 1

        workbook.close()


def make_note_in_offers(link):
    wb = openpyxl.load_workbook('config/data.xlsx')
    ws = wb['offers']
    row = ws.max_row + 1
    item = bot.parce(link, row-1)
    ws.cell(row, 1).value = row - 1
    for col in range(len(main_fields)):
        key = ws[1][col].value
        ws.cell(row, col + 1).value = str(item[key])
    print('Written ' + item['slug'])
    wb.save('config/data.xlsx')
    return row-1


def make_notes_in_child_offers(link, offer_key_id):
    wb = openpyxl.load_workbook('config/data.xlsx')
    ws = wb['child_offers']
    row = ws.max_row + 1
    items = bot.parce_child_offers(link, offer_key_id)
    for item in items:
        for col in range(len(child_fields)):
            key = ws[1][col].value
            ws.cell(row, col + 1).value = str(item[key])
        row += 1
        print(f'Written child offer for {item["agency"]}')
    wb.save('config/data.xlsx')